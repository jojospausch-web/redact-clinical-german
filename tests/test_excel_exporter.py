"""Tests for the Excel export utility."""

import io
import tempfile
from pathlib import Path

import fitz
import openpyxl
import pytest

from src.excel_exporter import export_to_excel, extract_text_from_pdf


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _create_simple_pdf(text: str) -> str:
    """Create a temporary single-page PDF with the given text.

    Returns the path; caller is responsible for deletion.
    """
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        path = f.name

    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    page.insert_text((50, 100), text, fontsize=11)
    doc.save(path)
    doc.close()
    return path


# ---------------------------------------------------------------------------
# export_to_excel tests
# ---------------------------------------------------------------------------

class TestExportToExcel:
    """Tests for export_to_excel()."""

    def test_header_row(self):
        """The first row must contain the column headers Dokument and Text."""
        buf = io.BytesIO()
        export_to_excel([], buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "Dokument"
        assert ws.cell(row=1, column=2).value == "Text"

    def test_single_document(self):
        """A single result entry is written to row 2."""
        buf = io.BytesIO()
        export_to_excel(
            [{"document": "patient_001.pdf", "text": "Anonymisierter Inhalt."}],
            buf,
        )
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "patient_001.pdf"
        assert ws.cell(row=2, column=2).value == "Anonymisierter Inhalt."

    def test_multiple_documents(self):
        """Each result entry occupies its own row."""
        results = [
            {"document": "doc_a.pdf", "text": "Text A"},
            {"document": "doc_b.pdf", "text": "Text B"},
            {"document": "doc_c.pdf", "text": "Text C"},
        ]
        buf = io.BytesIO()
        export_to_excel(results, buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        assert ws.max_row == 4  # 1 header + 3 data rows
        assert ws.cell(row=4, column=1).value == "doc_c.pdf"
        assert ws.cell(row=4, column=2).value == "Text C"

    def test_empty_results(self):
        """An empty results list produces only the header row."""
        buf = io.BytesIO()
        export_to_excel([], buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        assert ws.max_row == 1

    def test_multiline_text_stored_in_single_cell(self):
        """Multi-line text is stored verbatim in a single cell."""
        text = "Zeile 1\nZeile 2\nZeile 3"
        buf = io.BytesIO()
        export_to_excel([{"document": "multi.pdf", "text": text}], buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        assert ws.cell(row=2, column=2).value == text

    def test_save_to_file(self, tmp_path):
        """export_to_excel accepts a file path string and creates a valid .xlsx."""
        output = str(tmp_path / "output.xlsx")
        export_to_excel([{"document": "x.pdf", "text": "hello"}], output)
        assert Path(output).exists()
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "x.pdf"
        assert ws.cell(row=2, column=2).value == "hello"

    def test_worksheet_title(self):
        """The worksheet must be named 'Anonymisierte Texte'."""
        buf = io.BytesIO()
        export_to_excel([], buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        assert "Anonymisierte Texte" in wb.sheetnames


# ---------------------------------------------------------------------------
# extract_text_from_pdf tests
# ---------------------------------------------------------------------------

class TestExtractTextFromPdf:
    """Tests for extract_text_from_pdf()."""

    def test_extracts_text(self):
        """Text inserted into a PDF is returned by extract_text_from_pdf."""
        pdf_path = _create_simple_pdf("Hallo Welt")
        try:
            text = extract_text_from_pdf(pdf_path)
            assert "Hallo Welt" in text
        finally:
            Path(pdf_path).unlink(missing_ok=True)

    def test_empty_pdf_returns_empty_string(self):
        """A PDF with no text returns an empty or whitespace-only string."""
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            path = f.name
        try:
            doc = fitz.open()
            doc.new_page(width=595, height=842)
            doc.save(path)
            doc.close()
            text = extract_text_from_pdf(path)
            assert text.strip() == ""
        finally:
            Path(path).unlink(missing_ok=True)
